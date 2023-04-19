import openpyxl

# Access to workbook
#book = openpyxl.load_workbook("C://Users//gagan//OneDrive//Desktop//pythonExcelData.xlsx")
Dict = {} # Empty Dictionary
book = openpyxl.load_workbook("C://Users//gagan//PycharmProjects//PythonSelFramework//testData\pythonExcelData.xlsx")
sheet = book.active # Access to current sheet
cell = sheet.cell(row=1, column=2) # Access to specific cell
print(cell.value) # Get Cell text

sheet.cell(row=2, column=2).value = "Gaganpreet" # Writing value at row 2, column 2
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A3'].value)

for i in range(1, sheet.max_row+1): # To get rows
    if sheet.cell(row=i, column=1).value == "Testcase1":
        for j in range(2, sheet.max_column+1): # To get columns
        #if sheet.cell(row=i, column=j).value == "Pal": # only getting Pal from whole sheet
            #print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
