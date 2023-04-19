import openpyxl


class homePageData:

    test_homePage_data = [{"name":"Gaganpreet S Mann", "emailID":"gaganmann@gmail.com", "password":"password", "gender":"Female", "exampleText":"HelloAgain"},
                            {"name":"Parvinder Pal", "emailID":"parvinderpal@yahoo.com", "password":"password", "gender":"Female", "exampleText":"Hi"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}  # Empty Dictionary
        book = openpyxl.load_workbook("C://Users//gagan//PycharmProjects//PythonSelFramework//testData\pythonExcelData.xlsx")
        sheet = book.active  # Access to current sheet
        for i in range(1, sheet.max_row + 1):  # To get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # To get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]